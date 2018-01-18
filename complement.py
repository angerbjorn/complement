#!/usr/bin/python3

import sys
assert sys.version_info >= (3,0)
import optparse
import xml.etree.ElementTree
import random
from openpyxl import load_workbook, Workbook
from xml.dom.minidom import parseString

if __name__ == "__main__":
	parser = optparse.OptionParser(usage="Usage: %prog [OPTION]... <SPREADSHEET.xlsx>...", description="Generates a complementary .nessus v2 file from one or more excel spreadsheets.", epilog='Open Source MIT License. Written by Christian Angerbjorn')
	parser.add_option("-v", "--verbose", action="store_true")
	parser.add_option("-o", "--output-file", help='Save data to this .nessus file')
	parser.add_option("-t", "--make-template", help='Create an example excel template file, and exit.')
	(ops, args) = parser.parse_args()

	if len(args) == 0 and not ops.make_template:
		parser.error("At least one excel spreadsheet is required!")

	if not ops.make_template:
		severity = {'none': '0', 'low': '1', 'medium': '2', 'high': '3', 'critical': '4'}
		root = xml.etree.ElementTree.Element('NessusClientData_v2')
		report = xml.etree.ElementTree.SubElement(root, 'Report')
		for ef in args:
			wb = load_workbook( filename=ef)

			# ensure all required data is there
			required = set( ["risk_factor", '@pluginName', 'IP', 'description', 'solution', '@port', "@svc_name"])
			for i, row in enumerate(wb.active.iter_rows(max_row=1)):
				for cell in row:
					required.discard( cell.value )
			if len( required ) != 0:
				print( 'Fatal:   Columns %s are required in %s' %( ','.join( list(required)), ef), file=sys.stderr)
				exit()
		
			for i, row in enumerate(wb.active.iter_rows(min_row=2)):
				reportHost = xml.etree.ElementTree.SubElement( report, 'ReportHost') 	
				ReportItem = xml.etree.ElementTree.SubElement( reportHost, 'ReportItem', attrib={"pluginID": str(random.randrange(900000,999999)),  "pluginFamily": "Excel import", "protocol": "tcp", "svc_name": "excel", 'port': '1'})
				for cell in row:
					header = wb.active[ cell.column + '1'].value
					if not header:
						print( 'Fatal:   Missing header data at cell: %s1' %(cell.column), file=sys.stderr)
						exit()					

					# severity is taken from risk_factor, and set on ReportItem
					if header == 'risk_factor':
						if not cell.value:
							print( 'Fatal:   Data is required as risk_value at cell: %s%s (carefully delete any empty rows in excel)' %(cell.column, cell.row ), file=sys.stderr)
							exit()
						try:
							ReportItem.set('severity', severity[cell.value.lower()] )
						except KeyError as key:
							print( "Fatal:   risk_factor is %s and can only be either of none, low, medium, high, or critical" %key, file=sys.stderr)
							exit()
						cell.value = cell.value.capitalize()
					
					# IP should be set as attribute on ReportHost
					if header == 'IP':
						reportHost.set('name', cell.value )
					# Items with @ should be attributes, like "@svc_name": "general"
					elif header and header.startswith('@'):
						ReportItem.set( header[1:] , str(cell.value))
					else:
						xml.etree.ElementTree.SubElement( ReportItem, header ).text = str(cell.value)					

		outFile = None
		if ops.output_file:
			outFile = open( ops.output_file, 'w' )
		print( parseString(xml.etree.ElementTree.tostring( root )).toprettyxml(), file=outFile)
		if ops.output_file:
			outFile.close()

	if ops.make_template:
		templateHead =	[ "@pluginID", 							"risk_factor", 	'@pluginName',								'description',													'solution',						'IP', 				"plugin_output",											"@pluginFamily",	"@protocol",	'@port',	"@svc_name"	]
		templateData =	[ 999762,  								"High",			"Default administrative password",			'The default administrative password has not been changed.',	'Please change the password',	'192.168.2.100',	"URL: https://192.168.2.100/ Password used: admin/admin",	"Complementary findings", 	"tcp",			"443",		"www"		]
		templateData2 =	[ 999762,  								"High",			"Default administrative password",			'The default administrative password has not been changed.',	'Please change the password',	'192.168.2.120',	"URL: http://192.168.2.120/ Password used: admin/password",	"Complementary findings", 	"tcp",			"80",		"www"		]
		templateData3 =	[ str(random.randrange(900000,999999)),  "Critical",	"SQL Injection",							'SQL injection found on the login form',						'Use parameterized sql queries','192.168.2.120',	"Injected data: 'union select @@version,1,1,1 --",			"Complementary findings", 	"tcp",			"443",		"www"		]
		# new book
		excel_wb = Workbook()
		excel_file = excel_wb.active
		excel_file.title = "findings"
		excel_file.append( templateHead )
		excel_file.append( templateData )
		excel_file.append( templateData2 )
		excel_file.append( templateData3 )
		excel_wb.save( ops.make_template )
		
		